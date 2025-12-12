import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
import os
import csv
from collections import defaultdict
import re

import os

FILE_COMPTEUR = "compteur_visites.txt"

# Initialisation du fichier si absent
if not os.path.exists(FILE_COMPTEUR):
    with open(FILE_COMPTEUR, "w") as f:
        f.write("0")

# Lecture + incrémentation
with open(FILE_COMPTEUR, "r") as f:
    compteur = int(f.read().strip() or 0)
compteur += 1
with open(FILE_COMPTEUR, "w") as f:
    f.write(str(compteur))

# Affichage dans la barre latérale ou le header
import streamlit as st
st.sidebar.markdown(f"👁️ Nombre de bullteins générés : {compteur}")


# Chargement base INSEE + fréquence
FREQ_FILLE = defaultdict(int)
FREQ_GARCON = defaultdict(int)
PRENOM_CSV_PATH = os.path.join(os.path.dirname(__file__), "prenoms.csv")

with open(PRENOM_CSV_PATH, encoding="utf-8") as f:
    reader = csv.DictReader(f, delimiter=';')
    prenom_col = next((c for c in reader.fieldnames if c.lower() in {"preusuel", "prenoms", "prenom"}), None)
    sexe_col = next((c for c in reader.fieldnames if c.lower() == "sexe"), None)
    nombre_col = next((c for c in reader.fieldnames if c.lower() in {"nombre", "nombres", "nb"}), None)
    for row in reader:
        prenom = row[prenom_col].strip().capitalize()
        sexe = row[sexe_col]
        try:
            nb = int(row[nombre_col])
        except Exception:
            nb = 1
        if prenom.startswith("_") or not prenom:
            continue
        if sexe == "1":
            FREQ_GARCON[prenom] += nb
        elif sexe == "2":
            FREQ_FILLE[prenom] += nb

def extract_nom_prenom(nom_complet):
    """
    Retourne : nom (majuscules), prénom (tout le reste, minuscules/capitales classiques)
    Ex : 'FARES Adam' => nom: 'FARES', prénom: 'Adam'
    """
    if not nom_complet:
        return "", ""
    mots = str(nom_complet).split()
    nom_parts = [part for part in mots if part.isupper()]
    prenom_parts = [part for part in mots if not part.isupper()]
    nom = " ".join(nom_parts)
    prenom = " ".join(prenom_parts)
    return nom, prenom.capitalize()

def detect_genre_majoritaire(prenom):
    if not prenom:
        return "u"
    # Prend le premier prénom (utile en cas de prénoms composés)
    prenom_principal = prenom.split()[0].capitalize()
    nb_f = FREQ_FILLE.get(prenom_principal, 0)
    nb_m = FREQ_GARCON.get(prenom_principal, 0)
    if nb_f > nb_m:
        return "f"
    elif nb_m > nb_f:
        return "m"
    else:
        return "u"
    

def appreciation_bull(prenom, moyenne, genre_):
    if moyenne is None or moyenne == "Abs":
        return f"Absent ce trimestre."
    try:
        m = float(moyenne)
    except Exception:
        return f"Erreur sur la moyenne."

    serieux = "sérieuse" if genre_ == 'f' else "sérieux"
    applique = "appliquée" if genre_ == 'f' else "appliqué"
    implique = "impliquée" if genre_ == 'f' else "impliqué"
    brillant = "brillante" if genre_ == 'f' else "brillant"
    attentif = "attentive" if genre_ == 'f' else "attentif"
    consciencieux ="consciencieuse" if genre_ == 'f' else "consciencieux" 

    # 19.5 - 20
    if m >= 19.5:
        return f"Un trimestre remarquable pour {prenom}, dont le travail et l’engagement sont exemplaires. Félicitations."

    # 19 - 19.5
    if m >= 19:
        return f"Très bon trimestre pour {prenom}, élève très {brillant} et {implique} qui a fourni un travail exceptionnel. Félicitations."

    # 18.5 - 19
    if m >= 18.5:
        return f"Très beau trimestre pour {prenom}, dont la maîtrise et l’investissement méritent d’être salués. Le sérieux et la constance sont remarquables. Félicitations."

    # 18 - 18.5
    if m >= 18:
        return f"Très bon trimestre pour {prenom}, qui a été {serieux} et {consciencieux}. Son engagement et la qualité de son travail mettent en évidence une maîtrise remarquable. Félicitations."

    # 17.5 - 18
    if m >= 17.5:
        return f"{prenom} a fait preuve d’un très bon comportement et d’un travail très sérieux. Un trimestre solide et prometteur qui témoigne de ses capacités. Félicitations. "

    # 17 - 17.5
    if m >= 17:
        return f"Très bon trimestre pour {prenom}, élève {serieux} et {applique}. L’engagement est constant et la progression continue. Félicitations."

    # 16.5 - 17
    if m >= 16.5:
        return f"Un très bon trimestre pour {prenom}, dont le travail régulier et soigné porte bien ses fruits."

    # 16 - 16.5
    if m >= 16:
        return f"Bon trimestre pour {prenom}. Elève {serieux} et {attentif}. Les efforts fournis portent déjà leurs fruits et promettent de beaux progrès."

    # 15.5 - 16
    if m >= 15.5:
        return f"Bon trimestre pour {prenom}, dont le travail est appliqué et rigoureux. L’investissement reste encourageant."

    # 15 - 15.5
    if m >= 15:
        return f"Bon trimestre pour {prenom}, dont le travail est régulier et l'attitude est positive. Il faut maintenir cette dynamique positive."

    # 14.5 - 15
    if m >= 14.5:
        return f"Trimestre satisfaisant pour {prenom}, qui s’investit avec sérieux. Les acquis se consolident progressivement."

    # 14 - 14.5
    if m >= 14:
        return f"Bon ensemble pour {prenom}, malgré quelques irrégularités. Les acquis sont satisfaisants mais peuvent être encore consolidés."

    # 13.5 - 14
    if m >= 13.5:
        return f"Assez bon trimestre pour {prenom}. Le travail est sérieux et encourageant, mais en gagnant en régularité, les résultats seront meilleurs."

    # 13 - 13.5
    if m >= 13:
        return f"Assez bon trimestre pour {prenom}. Ses efforts sont encourageants et montrent une belle volonté de progresser ; il faut poursuive sur cette voie afin de gagner en confiance et en maîtrise."

    # 12.5 - 13
    if m >= 12.5:
        return f"Trimestre correct pour {prenom}. Un travail plus approfondi et constant permettrait d’atteindre un niveau supérieur."

    # 12 - 12.5
    if m >= 12:
        return f"Ensemble assez satisfaisant mais perfectible. {prenom} peut gagner en régularité pour renforcer ses acquis."

    # 11.5 - 12
    if m >= 11.5:
        return f"Résultats fragiles pour {prenom}. Avec un peu plus de régularité et d’attention, les progrès seront encore plus remarquables."

    # 11 - 11.5
    if m >= 11:
        return f"Résultats moyens. {prenom} doit gagner en constance et en concentration pour progresser."

    # 10.5 - 11
    if m >= 10.5:
        return f"Résultats encourageants pour {prenom}. Ses efforts sont prometteurs et, en poursuivant avec sérieux, elle consolidera encore davantage ses acquis."

    # 10 - 10.5
    if m >= 10:
        return f"Résultats fragiles pour {prenom}. Le travail est encore irrégulier, mais avec de la persévérance et du travail, des progrès sont possibles."

    # 9.5 - 10
    if m >= 9.5:
        return f"Résultats insuffisants. {prenom} doit renforcer son investissement pour éviter que les difficultés ne s’accentuent."

    # 9 - 9.5
    if m >= 9:
        return f"Résultats insuffisants. {prenom} doit travailler de manière plus régulière et structurée afin d'améliorer ses résultats."

    # 8.5 - 9
    if m >= 8.5:
        return f"Des résultats insuffisants et des difficultés persistantes pour {prenom}. Avec un travail régulier, sérieux et un accompagnement adapté, des progrès sont possibles."

    # 8 - 8.5
    if m >= 8:
        return f"Résultats insuffisants pour {prenom}. Plus de motivation et d’implication permettront une progression réelle."

    # 7.5 - 8
    if m >= 7.5:
        return f"Trimestre très insuffisant. {prenom} doit réagir rapidement et s’engager davantage dans le travail."

    # 7 - 7.5
    if m >= 7:
        return f"Résultats faibles. {prenom} doit s’investir sérieusement pour sortir de cette situation fragile."

    # 6.5 - 7
    if m >= 6.5:
        return f"Résultats très insuffisants pour {prenom}. Une reprise sérieuse et régulière du travail est indispensable."

    # 6 - 6.5
    if m >= 6:
        return f"Résultats très insuffisants. {prenom} doit réagir de toute urgence et adopter un rythme de travail plus soutenu."

    # 5 - 6
    if m >= 5:
        return f"Résultats très faibles, pénalisés par un manque de travail. {prenom} doit s’impliquer beaucoup plus sérieusement."

    # 3 - 5
    if m >= 3:
        return f"Résultats inquiétants pour {prenom}. Les difficultés sont importantes et nécessitent un suivi régulier et un travail approfondi."

    # 0 - 3
    if m > 0:
        return f"Résultats alarmants. {prenom} doit impérativement reprendre le travail avec sérieux et constance."

    return f"Données manquantes."


def generer_appreciations_excel_selection(input_path, output_path):
    wb_src = openpyxl.load_workbook(input_path)
    ws_src = wb_src.active
    entetes = [str(cell.value).strip().lower() if cell.value else "" for cell in ws_src[1]]
    col_nom = next((i for i, h in enumerate(entetes) if "élève" in h or "nom" in h), 0)
    col_moy = next((i for i, h in enumerate(entetes) if "moyenne" in h), 1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    ws.append(["Nom", "Prénom", "Moyenne", "Appréciation générale"])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=16)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for i, cell in enumerate(ws[1]):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    alt_fill = PatternFill(start_color="E8F0F7", end_color="E8F0F7", fill_type="solid")
    row_num = 2
    for src_row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row):
        nom_complet = src_row[col_nom].value
        moyenne = src_row[col_moy].value
        if nom_complet is None or str(nom_complet).strip() == "":
            continue
        nom, prenom = extract_nom_prenom(nom_complet)
        genre_ = detect_genre_majoritaire(prenom)
        appreciation = appreciation_bull(prenom, moyenne, genre_)
        ws.append([nom, prenom, moyenne, appreciation])

        for col_idx in range(1, 5):
            cell = ws.cell(row=row_num, column=col_idx)
            if row_num % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="left" if col_idx in [1,2] else "center", vertical="top", wrap_text=True)
            cell.font = Font(size=14)
            cell.border = border
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row_num += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 150
    ws.row_dimensions[1].height = 28
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 40

    wb.save(output_path)

st.set_page_config(page_title="Bulletin Automatique", page_icon="📝", layout="centered")
st.title("📝 Générateur d'appréciations pour bulletins scolaires")
st.header("Outil d'aide pour les professeurs")
st.markdown(
"""
Application créé par M. FARES (Professeur de mathématiques).

- Le programme fournit une appréciation en fonction de la moyenne générale.
- Les appréciations sont accordées selon le genre  (de données de l'INSEE).
"""
)

uploaded_file = st.file_uploader("📂 Dépose ton fichier Excel (.xlsx)", type=["xlsx"])

if uploaded_file is not None:
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_in:
            temp_in.write(uploaded_file.read())
            temp_in.flush()
            temp_out = temp_in.name.replace(".xlsx", "_appreciations.xlsx")
            generer_appreciations_excel_selection(temp_in.name, temp_out)
            st.success("✅ Fichier traité avec succès !")
            with open(temp_out, "rb") as f:
                st.download_button(
                    label="📥 Télécharger le bulletin avec appréciations",
                    data=f,
                    file_name="Bulletin_avec_appreciations.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            os.remove(temp_out)
        os.remove(temp_in.name)
    except Exception as e:
        st.error(f"❌ Erreur lors du traitement : {str(e)}")

st.markdown("""---
**💡 Astuces :**
- Extraire les notes de la classe à partir de PRONOTE et les insérer dans un fichier excel.
- Le fichier excel doit contenir le nom de l'élève en MAJUSCULE et le prénom en miniscule.
- Le fichier excel doit contenir sur la première colonne (NOM et prénom) et sur la deuxième colonne (La moyenne).
- L'appréciation prend en compte la moyenne de l'élève, c'est au professeur de l'adapter au profil de l'élève.
""")
